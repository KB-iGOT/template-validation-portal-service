import os
import time
from configparser import ConfigParser, ExtendedInterpolation
import xlrd
import uuid
import csv
from bson.objectid import ObjectId
import json
from datetime import datetime
import requests
from difflib import get_close_matches
from requests import post, get, delete
import sys
import time
import shutil
from xlutils.copy import copy
import shutil
import re
from xlrd import open_workbook
from xlutils.copy import copy as xl_copy
import logging.handlers
import time
from logging.handlers import TimedRotatingFileHandler
import xlsxwriter
import argparse
import sys
from os import path
import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles import colors
from openpyxl.cell import Cell
from backend.src.main.modules.config import *
# from common_config import *
import threading

millisecond = None
childsolutionid = ""
regex = "\"?([-a-zA-Z0-9.`?{}]+@\w+\.\w+)\"?"
class Helpers:
    def __init__(self):
        self.millisecond = None

    def programCreation(accessToken, parentFolder, externalId, pName, orgIds,creatorKeyCloakId, creatorName):
        global programIdForSuccess, programExternalId
       
        programCreationurl = internal_kong_ip_core +  programcreationurl
        
        programExternalId = externalId
        payload = json.dumps({
            "externalId": externalId,
            "name": pName,
            "description": "Program",
            "resourceType": [
                "program"
            ],
            "language": [
                "English"
            ],
            "keywords": [],
            "concepts": [],
            "createdFor": orgIds,
            "rootOrganisations": orgIds,
            "startDate": startDateOfProgram,
            "endDate": endDateOfProgram,
            "imageCompression": {
                "quality": 10
            },
            "creator": creatorName,
            "owner": creatorKeyCloakId,
            "author": creatorKeyCloakId,
            # "metaInformation": {
            #     # "state":entitiesPGM.split(","),
            #     # "role": mainRole.split(",")
            #     },
                "requestForPIIConsent":False
            })

        
        headers = {'X-authenticated-user-token': accessToken,
                   'internal-access-token': internal_access_token,
                   'Content-Type': 'application/json',
                   'Authorization':authorization}

        # program creation 
        responsePgmCreate = requests.request("POST", programCreationurl, headers=headers, data=(payload))
        fileheader = [pName, ('Program Sheet Validation'), ('Passed')]
        if responsePgmCreate.status_code == 200:
            responsePgmCreateResp = responsePgmCreate.json()
            programIdForSuccess = responsePgmCreateResp['result']["_id"]
        else:
            return None

    def createFileStructForProgram(programFile):
        if not os.path.isdir('resourceFile'):
            os.mkdir('resourceFile')
        if "/" in str(programFile):
            fileNameSplit = str(programFile).split('/')[-1:]
        else :

            fileNameSplit = os.path.basename(programFile)

        if isinstance(fileNameSplit, list):
            fileNameSplit = fileNameSplit[0]
        # fileNameSplit = str(programFile)
        if fileNameSplit.endswith(".xlsx"):
            ts = str(time.time()).replace(".", "_")
            
            folderName = fileNameSplit.replace(".xlsx", "-" + str(ts))
            os.mkdir('resourceFile/' + str(folderName))
            path = os.path.join('resourceFile', str(folderName))
        else:
            return None
        returnPathStr = os.path.join('resourceFile', str(folderName))

        return returnPathStr

# Function create File structure for Solutions
    def createFileStruct(MainFilePath, addSolutionFile):
        if not os.path.isdir(MainFilePath + '/SolutionFiles'):
            os.mkdir(MainFilePath + '/SolutionFiles')
        if "\\" in str(addSolutionFile):
            fileNameSplit = str(addSolutionFile).split('\\')[-1:]
        elif "/" in str(addSolutionFile):
            fileNameSplit = str(addSolutionFile).split('/')[-1:]
        else:
            fileNameSplit = str(addSolutionFile)
        if ".xlsx" in str(fileNameSplit[0]):
            ts = str(time.time()).replace(".", "_")
            folderName = fileNameSplit[0].replace(".xlsx", "-" + str(ts))
            os.mkdir(MainFilePath + '/SolutionFiles/' + str(folderName))
            path = os.path.join(MainFilePath + '/SolutionFiles', str(folderName))
            path = os.path.join(path, str('apiHitLogs'))
            os.mkdir(path)
        else:
            return None
        returnPathStr = os.path.join(MainFilePath + '/SolutionFiles', str(folderName))

        if not os.path.isdir(returnPathStr + "/user_input_file"):
            os.mkdir(returnPathStr + "/user_input_file")

        shutil.copy(addSolutionFile, os.path.join(returnPathStr + "user_input_file.xlsx"))
        # shutil.copy(programFile, os.path.join(returnPathStr + "user_input_file"))
        return returnPathStr
    

    # Generate access token for the APIs. 
    def generateAccessToken(solutionName_for_folder_path):
    # production search user api - start
        headerKeyClockUser = {'Content-Type': keyclockapicontent_type}
    
        responseKeyClockUser = requests.post(url=host + (keyclockapiurl), headers=headerKeyClockUser,
                                         data=(keyclockapibody))
        if responseKeyClockUser.status_code == 200:
            responseKeyClockUser = responseKeyClockUser.json()
            accessTokenUser = responseKeyClockUser['access_token']

        return accessTokenUser
    
    def getProgramInfo(accessTokenUser, solutionName_for_folder_path, solutionNameInp):
        global programID, programExternalId, programDescription, isProgramnamePresent, programName, programIdForSuccess
        programName = solutionNameInp
        programUrl = internal_kong_ip_core + fetchprograminfoapiurl + solutionNameInp.lstrip().rstrip()

        headersProgramSearch = {'Authorization': authorization,
                                'Content-Type': 'application/json', 
                                'X-authenticated-user-token':accessTokenUser,
                                'internal-access-token': internal_access_token}
        responseProgramSearch = requests.post(url=programUrl, headers=headersProgramSearch)
       
        if responseProgramSearch.status_code == 200:
            responseProgramSearch = responseProgramSearch.json()
            countOfPrograms = len(responseProgramSearch['result']['data'])
            if countOfPrograms == 0:
                return False
            else:
                getProgramDetails = []
                for eachPgm in responseProgramSearch['result']['data']:
                    if eachPgm['isAPrivateProgram'] == False:
                        programID = eachPgm['_id']
                        programExternalId = eachPgm['externalId']
                        programDescription = eachPgm['description']
                        isAPrivateProgram = eachPgm['isAPrivateProgram']
                        getProgramDetails.append([programID, programExternalId, programDescription, isAPrivateProgram])
                        if len(getProgramDetails) == 0:
                            print("Total " + str(len(getProgramDetails)) + " backend programs found with the name : " + programName.lstrip().rstrip())
                            sys.exit("Aborting...")
                        elif len(getProgramDetails) > 1:
                            print("Total " + str(len(getProgramDetails)) + " backend programs found with the name : " + programName.lstrip().rstrip())
                            sys.exit("Aborting...")

                        else:
                            programID = getProgramDetails[0][0]
                            programIdForSuccess = getProgramDetails[0][0]
                            programExternalId = getProgramDetails[0][1]
                            programDescription = getProgramDetails[0][2]
                            isAPrivateProgram = getProgramDetails[0][3]
                            isProgramnamePresent = True

        else:
            return None

        return True
    def checkEmailValidation(email):
        if (re.search(regex, email)):
            return True
        else:
            return False
    def fetchUserDetails(accessToken, KBId):
        global OrgName,creatorId
        url =  host + userinfoapiurl
        headers = {'Content-Type': 'application/json',
               'Authorization': authorizationforhost}
        isEmail = Helpers.checkEmailValidation(KBId.lstrip().rstrip())
        
        if isEmail:
            body = "{\n  \"request\": {\n    \"filters\": {\n    \t\"email\": \"" + KBId.lstrip().rstrip() + "\"\n    },\n      \"fields\" :[],\n    \"limit\": 1000,\n    \"sort_by\": {\"createdDate\": \"desc\"}\n  }\n}"
        else:
            body = "{\n  \"request\": {\n    \"filters\": {\n    \t\"userName\": \"" + KBId.lstrip().rstrip() + "\"\n    },\n      \"fields\" :[],\n    \"limit\": 1000,\n    \"sort_by\": {\"createdDate\": \"desc\"}\n  }\n}"

        responseUserSearch = requests.request("POST", url, headers=headers, data=body)
        response_json = responseUserSearch.json()

        if responseUserSearch.status_code == 200:
            responseUserSearch = responseUserSearch.json()
            if responseUserSearch['result']['response']['content']:
                userKeycloak = responseUserSearch['result']['response']['content'][0]['userId']
                creatorId = userKeycloak
                userName = responseUserSearch['result']['response']['content'][0]['userName']
                firstName = responseUserSearch['result']['response']['content'][0]['firstName']
                rootOrgId = responseUserSearch['result']['response']['content'][0]['rootOrgId']
                for index in responseUserSearch['result']['response']['content'][0]['organisations']:
                    if rootOrgId == index['organisationId']:
                        roledetails = index['roles']
                        rootOrgName = index['orgName']
                        # OrgName.append(index['orgName'])
            else:
                return("-->Given username/email is not present in KB platform<--.")
        else:
            return(responseUserSearch.text)

        return [userKeycloak, userName, firstName,roledetails,rootOrgName,rootOrgId]
    
    def SolutionFileCheck(filePathAddPgm, accessToken, parentFolder, MainFilePath):
        global creatorId,solutionNameForSuccess
        wbPgm = xlrd.open_workbook(filePathAddPgm, on_demand=True)
        global solutionNameInp
        sheetNames = wbPgm.sheet_names()
        for sheetEnv in sheetNames:
            if sheetEnv.strip().lower() == 'details':
                detailsEnvSheet = wbPgm.sheet_by_name(sheetEnv)
                keysEnv = [detailsEnvSheet.cell(1, col_index_env).value for col_index_env in
                            range(detailsEnvSheet.ncols)]
                for row_index_env in range(2, detailsEnvSheet.nrows):
                    dictDetailsEnv = {keysEnv[col_index_env]: detailsEnvSheet.cell(row_index_env, col_index_env).value
                                        for
                                        col_index_env in range(detailsEnvSheet.ncols)}
                    solutionNameInp = dictDetailsEnv['solution_name'].encode('utf-8').decode('utf-8')
                    print(solutionNameInp)
                    solutionNameForSuccess = solutionNameInp
                    global entitiesPGM

                    global startDateOfProgram, endDateOfProgram
                    startDateOfProgram = dictDetailsEnv['start_date']
                    endDateOfProgram = dictDetailsEnv['end_date']

                    # taking the start date of program from program template and converting YYYY-MM-DD 00:00:00 format

                    startDateArr = str(startDateOfProgram).split("-")
                    startDateOfProgram = startDateArr[2] + "-" + startDateArr[1] + "-" + startDateArr[0] + " 00:00:00"

                    # taking the end date of program from program template and converting YYYY-MM-DD 00:00:00 format

                    endDateArr = str(endDateOfProgram).split("-")
                    endDateOfProgram = endDateArr[2] + "-" + endDateArr[1] + "-" + endDateArr[0] + " 23:59:59"
                    if not Helpers.getProgramInfo(accessToken, parentFolder, solutionNameInp.encode('utf-8').decode('utf-8')):
                        extIdPGM = dictDetailsEnv['solution_name'].encode('utf-8').decode('utf-8')
                        programName = extIdPGM = dictDetailsEnv['solution_name'].encode('utf-8').decode('utf-8')
                        userDetails = Helpers.fetchUserDetails(accessToken, dictDetailsEnv['creator_username'])
                        OrgName=userDetails[4]
                        orgIds=Helpers.fetchOrgId(accessToken, parentFolder, OrgName)
                        creatorKeyCloakId = userDetails[0]
                        creatorName = userDetails[2]
                        if Helpers.getProgramInfo(accessToken, parentFolder, extIdPGM):
                            print("Program Created SuccessFully.")
                        else :
                            print("program creation API called")
                            Helpers.programCreation(accessToken, parentFolder, extIdPGM, programName,orgIds,creatorKeyCloakId, creatorName)



# fetch org Ids 
    def fetchOrgId(accessToken, parentFolder, OrgName):
        url = host + fetchorgdetails
        headers = {'Content-Type': 'application/json',
                   'Authorization': authorization,
                   'x-authenticated-user-token': accessToken}
        orgIds = []
        organisations = str(OrgName).split(",")
        for org in organisations:
            orgBody = {"id": "",
                       "ts": "",
                       "params": {
                           "msgid": "",
                           "resmsgid": "",
                           "status": "success"
                       },
                       "request": {
                           "filters": {
                               "orgName": str(org).strip()
                           }
                       }}

            responseOrgSearch = requests.request("POST", url, headers=headers, data=json.dumps(orgBody))
            if responseOrgSearch.status_code == 200:
                responseOrgSearch = responseOrgSearch.json()
                if responseOrgSearch['result']['response']['content']:
                    orgId = responseOrgSearch['result']['response']['content'][0]['id']
                    orgIds.append(orgId)
                else:
                    return("Email is not present in KB")
            else:
               return(responseOrgSearch.text)
                
        return orgIds

    def solutionUpdate(solutionName_for_folder_path, accessToken, solutionId, bodySolutionUpdate):
        solutionUpdateApi = internal_kong_ip_core + solutionupdateapi + str(solutionId)
        headerUpdateSolutionApi = {
            'Content-Type': 'application/json',
            'Authorization': authorization,
            'X-authenticated-user-token': accessToken,
            'X-Channel-id': x_channel_id,
            "internal-access-token": internal_access_token
            }
        responseUpdateSolutionApi = requests.post(url=solutionUpdateApi, headers=headerUpdateSolutionApi,data=json.dumps(bodySolutionUpdate))
        
        if responseUpdateSolutionApi.status_code == 200:
            return True
        else:
            return False

    def createSurveySolution(parentFolder, wbSurvey, accessToken):
    
        sheetNames1 = wbSurvey.sheet_names()
        for sheetEnv in sheetNames1:
            if sheetEnv.strip().lower() == 'details':
                surveySolutionCreationReqBody = {}
                detailsEnvSheet = wbSurvey.sheet_by_name(sheetEnv)
                keysEnv = [detailsEnvSheet.cell(1, col_index_env).value for col_index_env in
                       range(detailsEnvSheet.ncols)]

                for row_index_env in range(2, detailsEnvSheet.nrows):
                    dictDetailsEnv = {keysEnv[col_index_env]: detailsEnvSheet.cell(row_index_env, col_index_env).value
                                  for
                                  col_index_env in range(detailsEnvSheet.ncols)}
                    surveySolutionCreationReqBody['name'] = dictDetailsEnv['solution_name'].encode('utf-8').decode('utf-8')
                    surveySolutionCreationReqBody["description"] = "survey Solution"
                    surveySolutionExternalId = str(uuid.uuid1())
                    surveySolutionCreationReqBody["externalId"] = surveySolutionExternalId
                    # if dictDetailsEnv['creator_username'].encode('utf-8').decode('utf-8') == "":
                    #     exceptionHandlingFlag = True
                    #     print('survey_creator_username column should not be empty in the details sheet')
                    #     sys.exit()
                    # else:
                    #     surveySolutionCreationReqBody['creator'] = dictDetailsEnv['Name_of_the_creator'].encode('utf-8').decode('utf-8')

                    
                    userDetails = Helpers.fetchUserDetails(accessToken, dictDetailsEnv['creator_username'])
                    # print(userDetails)
                    surveySolutionCreationReqBody['author'] = userDetails[0]
                    # print("surveySolutionCreationReqBody",surveySolutionCreationReqBody)

                    # Below script will convert date DD-MM-YYYY TO YYYY-MM-DD 00:00:00 to match the code syntax 

                    if dictDetailsEnv["start_date"]:
                        if type(dictDetailsEnv["start_date"]) == str:
                            startDateArr = None
                            startDateArr = (dictDetailsEnv["start_date"]).split("-")
                            surveySolutionCreationReqBody["startDate"] = startDateArr[2] + "-" + startDateArr[1] + "-" + \
                                                                     startDateArr[0] + " 00:00:00"
                        elif type(dictDetailsEnv["start_date"]) == float:
                            surveySolutionCreationReqBody["startDate"] = (
                            xlrd.xldate.xldate_as_datetime(dictDetailsEnv["start_date"],
                                                           wbSurvey.datemode)).strftime("%Y/%m/%d")
                        else:
                            surveySolutionCreationReqBody["startDate"] = ""
                        if dictDetailsEnv["end_date"]:
                            if type(dictDetailsEnv["end_date"]) == str:

                                endDateArr = None
                                endDateArr = (dictDetailsEnv["end_date"]).split("-")
                                surveySolutionCreationReqBody["endDate"] = endDateArr[2] + "-" + endDateArr[1] + "-" + \
                                                                       endDateArr[0] + " 23:59:59"
                            elif type(dictDetailsEnv["end_date"]) == float:
                                surveySolutionCreationReqBody["endDate"] = (
                                    xlrd.xldate.xldate_as_datetime(dictDetailsEnv["end_date"],
                                                               wbSurvey.datemode)).strftime("%Y/%m/%d")
                            else:
                                surveySolutionCreationReqBody["endDate"] = ""
                            enDt = surveySolutionCreationReqBody["endDate"]
                        
                            urlCreateSolutionApi =internal_kong_ip_survey+ surveysolutioncreationapiurl
                            headerCreateSolutionApi = {
                            'Content-Type': 'application/json',
                            'Authorization': authorization,
                            'X-authenticated-user-token': accessToken,
                            'X-Channel-id': x_channel_id,
                            'appName': appname
                        }
                            # print(surveySolutionCreationReqBody)
                            # sys.exit()
                            responseCreateSolutionApi = requests.post(url=urlCreateSolutionApi,
                                                                  headers=headerCreateSolutionApi,
                                                                  data=json.dumps(surveySolutionCreationReqBody))
                        
                            if responseCreateSolutionApi.status_code == 200:
                                responseCreateSolutionApi = responseCreateSolutionApi.json()
                                urlSearchSolution = internal_kong_ip_core + fetchsolutiondetails + "survey&page=1&limit=10&search=" + str(surveySolutionExternalId)
                                responseSearchSolution = requests.request("POST", urlSearchSolution,
                                                                      headers=headerCreateSolutionApi)
                            
                                if responseSearchSolution.status_code == 200:
                                    responseSearchSolutionApi = responseSearchSolution.json()
                                    surveySolutionExternalId = None
                                    surveySolutionExternalId = responseSearchSolutionApi['result']['data'][0]['externalId']
                                else:
                                    return("URL : " + urlSearchSolution)

                                solutionId = None
                                solutionId = responseCreateSolutionApi["result"]["solutionId"]
                                bodySolutionUpdate = {"creator": userDetails[2]}

                                return [solutionId, surveySolutionExternalId]
                            
                            else:
                                return("something went wrong here ")


    def uploadSurveyQuestions(parentFolder, wbSurvey, addSolutionFile, accessToken, surveySolutionExternalId, surveyParentSolutionId,millisecond):
        global childsolutionid
        sheetNam = wbSurvey.sheet_names()
        stDt = None
        enDt = None
        shCnt = 0
        for i in sheetNam:
            if i.strip().lower() == 'questions':
                sheetNam1 = wbSurvey.sheets()[shCnt]
            shCnt = shCnt + 1
        dataSort = [sheetNam1.row_values(i) for i in range(sheetNam1.nrows)]
        labels = dataSort[1]
        dataSort = dataSort[2:]
        dataSort.sort(key=lambda x: int(x[0]))
        openWorkBookSort1 = xl_copy(wbSurvey)
        sheet1 = openWorkBookSort1.add_sheet('questions_sequence_sorted')

        for idx, label in enumerate(labels):
            sheet1.write(0, idx, label)

        for idx_r, row in enumerate(dataSort):
            for idx_c, value in enumerate(row):
                sheet1.write(idx_r + 1, idx_c, value)
        newFileName = str(addSolutionFile)
        openWorkBookSort1.save(newFileName)
        openNewFile = xlrd.open_workbook(newFileName, on_demand=True)
        wbSurvey = openNewFile
        sheetNames = wbSurvey.sheet_names()
        for sheet2 in sheetNames:
            if sheet2.strip().lower() == 'questions_sequence_sorted':
                questionsList = []
                questionsSheet = wbSurvey.sheet_by_name(sheet2.lower())
                keys2 = [questionsSheet.cell(0, col_index2).value for col_index2 in
                         range(questionsSheet.ncols)]
                for row_index2 in range(1, questionsSheet.nrows):
                    d2 = {keys2[col_index2]: questionsSheet.cell(row_index2, col_index2).value
                          for col_index2 in range(questionsSheet.ncols)}
                    questionsList.append(d2)
                questionSeqByEcmArr = []
                quesSeqCnt = 1.0
                questionUploadFieldnames = []
                questionUploadFieldnames = ['solutionId', 'instanceParentQuestionId','hasAParentQuestion', 'parentQuestionOperator','parentQuestionValue', 'parentQuestionId','externalId', 'question0', 'question1', 'tip','hint', 'instanceIdentifier', 'responseType','dateFormat', 'autoCapture', 'validation','validationIsNumber', 'validationRegex','validationMax', 'validationMin', 'file','fileIsRequired', 'fileUploadType','allowAudioRecording', 'minFileCount','maxFileCount', 'caption', 'questionGroup','modeOfCollection', 'accessibility', 'showRemarks','rubricLevel', 'isAGeneralQuestion', 'R1','R1-hint', 'R2', 'R2-hint', 'R3', 'R3-hint', 'R4','R4-hint', 'R5', 'R5-hint', 'R6', 'R6-hint', 'R7','R7-hint', 'R8', 'R8-hint', 'R9', 'R9-hint', 'R10','R10-hint', 'R11', 'R11-hint', 'R12', 'R12-hint','R13', 'R13-hint', 'R14', 'R14-hint', 'R15','R15-hint', 'R16', 'R16-hint', 'R17', 'R17-hint','R18', 'R18-hint', 'R19', 'R19-hint', 'R20','R20-hint', 'sectionHeader', 'page','questionNumber', '_arrayFields']
                # print(questionsList,"12333321")
                ts = str(time.time())
                for ques in questionsList:
                    # print(ques,"quweeeee")
                    # sys.exit()

                    questionFilePath = parentFolder + '/questionUpload/'
                    file_exists_ques = os.path.isfile(
                        parentFolder + '/questionUpload/uploadSheet.csv')
                    if not os.path.exists(questionFilePath):
                        os.mkdir(questionFilePath)
                    with open(parentFolder + '/questionUpload/uploadSheet.csv', 'a',
                              encoding='utf-8') as questionUploadFile:
                        writerQuestionUpload = csv.DictWriter(questionUploadFile, fieldnames=questionUploadFieldnames,lineterminator='\n')
                        if not file_exists_ques:
                            writerQuestionUpload.writeheader()
                        questionFileObj = {}
                        surveyExternalId = None
                        questionFileObj['solutionId'] = surveySolutionExternalId
                        if ques['instance_parent_question_id'].encode('utf-8').decode('utf-8'):
                            questionFileObj['instanceParentQuestionId'] = ques[
                                                                              'instance_parent_question_id'].strip() + '_' + str(ts)
                        else:
                            questionFileObj['instanceParentQuestionId'] = 'NA'
                        if ques['parent_question_id'].encode('utf-8').decode('utf-8').strip():
                            questionFileObj['hasAParentQuestion'] = 'YES'
                            if ques['show_when_parent_question_value_is'] == 'or':
                                questionFileObj['parentQuestionOperator'] = '||'
                            else:
                                questionFileObj['parentQuestionOperator'] = ques['show_when_parent_question_value_is']
                            if type(ques['parent_question_value']) != str:
                                if (ques['parent_question_value'] and ques[
                                    'parent_question_value'].is_integer() == True):
                                    questionFileObj['parentQuestionValue'] = int(ques['parent_question_value'])
                                elif (ques['parent_question_value'] and ques[
                                    'parent_question_value'].is_integer() == False):
                                    questionFileObj['parentQuestionValue'] = ques['parent_question_value']
                            else:
                                questionFileObj['parentQuestionValue'] = ques['parent_question_value']
                                questionFileObj['parentQuestionId'] = ques['parent_question_id'].encode('utf-8').decode('utf-8').strip()+ '_' + str(ts)
                        else:
                            questionFileObj['hasAParentQuestion'] = 'NO'
                            questionFileObj['parentQuestionOperator'] = None
                            questionFileObj['parentQuestionValue'] = None
                            questionFileObj['parentQuestionId'] = None
                        questionFileObj['externalId'] = ques['question_id'].strip()+ '_' + str(ts)
                        if quesSeqCnt == ques['question_sequence']:
                            questionSeqByEcmArr.append(ques['question_id'].strip() + '_' + str(ts))
                            quesSeqCnt = quesSeqCnt + 1.0
                        if ques['question_language1']:
                            questionFileObj['question0'] = ques['question_language1']
                        else:
                            questionFileObj['question0'] = None
                        if ques['question_language2']:
                            questionFileObj['question1'] = ques['question_language2'].encode('utf-8').decode('utf-8')
                        else:
                            questionFileObj['question1'] = None
                        if ques['question_tip']:
                            questionFileObj['tip'] = ques['question_tip'].encode('utf-8').decode('utf-8')
                        else:
                            questionFileObj['tip'] = None
                        if ques['question_hint']:
                            questionFileObj['hint'] = ques['question_hint'].encode('utf-8').decode('utf-8')
                        else:
                            questionFileObj['hint'] = None
                        if ques['instance_identifier']:
                            questionFileObj['instanceIdentifier'] = ques['instance_identifier'].encode('utf-8').decode('utf-8')
                        else:
                            questionFileObj['instanceIdentifier'] = None
                        if ques['question_response_type'].strip().lower():
                            questionFileObj['responseType'] = ques['question_response_type'].strip().lower()
                        if ques['question_response_type'].strip().lower() == 'date':
                            questionFileObj['dateFormat'] = "DD-MM-YYYY"
                        else:
                            questionFileObj['dateFormat'] = None
                        if ques['question_response_type'].strip().lower() == 'date':
                            if ques['date_auto_capture'] and ques['date_auto_capture'] == 1:
                                questionFileObj['autoCapture'] = 'TRUE'
                            elif ques['date_auto_capture'] and ques['date_auto_capture'] == 0:
                                questionFileObj['autoCapture'] = 'false'
                            else:
                                questionFileObj['autoCapture'] = 'false'
                        else:
                            questionFileObj['autoCapture'] = None
                        if ques['response_required']:
                            if ques['response_required'] == 1:
                                questionFileObj['validation'] = 'TRUE'
                            elif ques['response_required'] == 0:
                                questionFileObj['validation'] = 'FALSE'
                        else:
                            questionFileObj['validation'] = 'FALSE'
                        if ques['question_response_type'].strip().lower() == 'number':
                            questionFileObj['validationIsNumber'] = 'TRUE'
                            questionFileObj['validationRegex'] = 'isNumber'
                            if (ques['max_number_value'] and ques['max_number_value'].is_integer() == True):
                                questionFileObj['validationMax'] = int(ques['max_number_value'])
                            elif (ques['max_number_value'] and ques['max_number_value'].is_integer() == False):
                                questionFileObj['validationMax'] = ques['max_number_value']
                            else:
                                questionFileObj['validationMax'] = 10000

                            if (ques['min_number_value'] and ques['min_number_value'].is_integer() == True):
                                questionFileObj['validationMin'] = int(ques['min_number_value'])
                            elif (ques['min_number_value'] and ques['min_number_value'].is_integer() == False):
                                questionFileObj['validationMin'] = ques['min_number_value']
                            else:
                                questionFileObj['validationMax'] = 10000

                            if (ques['min_number_value'] and ques['min_number_value'].is_integer() == True):
                                questionFileObj['validationMin'] = int(ques['min_number_value'])
                            elif (ques['min_number_value'] and ques['min_number_value'].is_integer() == False):
                                questionFileObj['validationMin'] = ques['min_number_value']
                            else:
                                questionFileObj['validationMin'] = 0
                        if ques['question_response_type'].strip().lower() == 'text':
                            if (ques['question_response_validation'] and ques['question_response_validation'].strip().lower() == "email"):
                                questionFileObj['validationRegex'] = "^[\w-\.]+@([\w-]+\.)+[\w-]{2,4}$"
                            elif (ques['question_response_validation'] and ques['question_response_validation'].strip().lower() == "text without special char"):
                                questionFileObj['validationRegex'] = "^[a-zA-Z ]+$"
                            elif (ques['question_response_validation'] and ques['question_response_validation'].strip().lower() == "text with special char"):
                                questionFileObj['validationRegex'] = "^[a-zA-Z0-9\s!@#$%^&*()_+{}\[\]:;<>,.?/~`|-]*$"

                        elif ques['question_response_type'].strip().lower() == 'slider':
                            questionFileObj['validationIsNumber'] = None
                            questionFileObj['validationRegex'] = 'isNumber'
                            if (ques['max_number_value'] and ques['max_number_value'].is_integer() == True):
                                questionFileObj['validationMax'] = int(ques['max_number_value'])
                            elif (ques['max_number_value'] and ques['max_number_value'].is_integer() == False):
                                questionFileObj['validationMax'] = ques['max_number_value']
                            else:
                                questionFileObj['validationMax'] = 5

                            if (ques['min_number_value'] and ques['min_number_value'].is_integer() == True):
                                questionFileObj['validationMin'] = int(ques['min_number_value'])
                            elif (ques['min_number_value'] and ques['min_number_value'].is_integer() == False):
                                questionFileObj['validationMin'] = ques['min_number_value']
                            else:
                                questionFileObj['validationMin'] = 0
                        else:
                            questionFileObj['validationIsNumber'] = None
                            questionFileObj['validationRegex'] = None
                            questionFileObj['validationMax'] = None
                            questionFileObj['validationMin'] = None
                        if ques['file_upload'] == 1:
                            questionFileObj['file'] = 'Snapshot'
                            questionFileObj['fileIsRequired'] = 'TRUE'
                            questionFileObj['fileUploadType'] = 'png,image/png,jpg,image/jpg,heic,image/heic,heif,image/heif,hevc,image/hevc,jpeg,image/jpeg,webp,image/webp,mp4,video/mp4,webm,video/webm,mkv,video/mkv,avi,video/avi,wmv,video/wmv,flv,video/flv,3GP,video/3GP,ogg,video/ogg,m4v,video/m4v,mov,video/mov,mp3,audio/mp3,wav,audio/wav,mpeg,audio/mpeg,pdf,application/pdf,xlsx,application/xlsx,txt,application/txt,doc,application/doc,docx,application/docx,xls,application/xls,csv,application/csv'
                            questionFileObj['minFileCount'] = 0
                            questionFileObj['maxFileCount'] = 10
                        elif ques['file_upload'] == 0:
                            questionFileObj['file'] = 'NA'
                            questionFileObj['fileIsRequired'] = None
                            questionFileObj['fileUploadType'] = None
                            questionFileObj['minFileCount'] = None
                            questionFileObj['maxFileCount'] = None

                        questionFileObj['caption'] = 'FALSE'
                        questionFileObj['questionGroup'] = 'A1'
                        questionFileObj['modeOfCollection'] = 'onfield'
                        questionFileObj['accessibility'] = 'No'
                        if ques['show_remarks'] == 1:
                            questionFileObj['showRemarks'] = 'TRUE'
                        elif ques['show_remarks'] == 0:
                            questionFileObj['showRemarks'] = 'FALSE'
                        questionFileObj['rubricLevel'] = None
                        questionFileObj['isAGeneralQuestion'] = None
                        if ques['question_response_type'].strip().lower() == 'radio' or ques[
                            'question_response_type'].strip() == 'multiselect':
                            for quesIndex in range(1, 21):
                                if type(ques['response(R' + str(quesIndex) + ')']) != str:
                                    if (ques['response(R' + str(quesIndex) + ')'] and ques[
                                        'response(R' + str(quesIndex) + ')'].is_integer() == True):
                                        questionFileObj['R' + str(quesIndex) + ''] = int(
                                            ques['response(R' + str(quesIndex) + ')'])
                                    elif (ques['response(R' + str(quesIndex) + ')'] and ques[
                                        'response(R' + str(quesIndex) + ')'].is_integer() == False):
                                        questionFileObj['R' + str(quesIndex) + ''] = ques[
                                            'response(R' + str(quesIndex) + ')']
                                else:
                                    questionFileObj['R' + str(quesIndex) + ''] = ques[
                                        'response(R' + str(quesIndex) + ')']

                                if type(ques['response(R' + str(quesIndex) + ')_hint']) != str:
                                    if (ques['response(R' + str(quesIndex) + ')_hint'] and ques[
                                        'response(R' + str(quesIndex) + ')_hint'].is_integer() == True):
                                        questionFileObj['R' + str(quesIndex) + '-hint'] = int(
                                            ques['response(R' + str(quesIndex) + ')_hint'])
                                    elif (ques['response(R' + str(quesIndex) + ')_hint'] and ques[
                                        'response(R' + str(quesIndex) + ')_hint'].is_integer() == False):
                                        questionFileObj['R' + str(quesIndex) + '-hint'] = ques[
                                            'response(R' + str(quesIndex) + ')_hint']
                                else:
                                    questionFileObj['R' + str(quesIndex) + '-hint'] = ques[
                                        'response(R' + str(quesIndex) + ')_hint']
                                questionFileObj['_arrayFields'] = 'parentQuestionValue'
                        else:
                            for quesIndex in range(1, 21):
                                questionFileObj['R' + str(quesIndex)] = None
                                questionFileObj['R' + str(quesIndex) + '-hint'] = None
                        if ques['section_header'].encode('utf-8').decode('utf-8'):
                            questionFileObj['sectionHeader'] = ques['section_header']
                        else:
                            questionFileObj['sectionHeader'] = None

                        questionFileObj['page'] = ques['page']
                        if type(ques['question_number']) != str:
                            if ques['question_number'] and ques['question_number'].is_integer() == True:
                                questionFileObj['questionNumber'] = int(ques['question_number'])
                            elif ques['question_number']:
                                questionFileObj['questionNumber'] = ques['question_number']
                            else:
                                questionFileObj['questionNumber'] = ques['question_number']
                        writerQuestionUpload.writerow(questionFileObj)
                urlQuestionsUploadApi = internal_kong_ip_survey + questionuploadapiurl
                # print("urlQuestionsUploadApi:",urlQuestionsUploadApi)
                headerQuestionUploadApi = {
                    'Authorization': authorization,
                    'X-authenticated-user-token': accessToken,
                    'X-Channel-id': x_channel_id
                }
                
                filesQuestion = {
                    'questions': open(parentFolder + '/questionUpload/uploadSheet.csv', 'rb')
                }
                
                responseQuestionUploadApi = requests.post(url=urlQuestionsUploadApi,
                                                          headers=headerQuestionUploadApi, files=filesQuestion)
                if responseQuestionUploadApi.status_code == 200:

                    with open(parentFolder + '/questionUpload/uploadInternalIdsSheet.csv', 'w+',encoding='utf-8') as questionRes:
                        questionRes.write(responseQuestionUploadApi.text)
                    urlImportSoluTemplate = internal_kong_ip_survey + importsurveysolutiontemplateurl + str(surveyParentSolutionId) + "?appName=manage-learn"
                    headerImportSoluTemplateApi = {
                        'Authorization': authorization,
                        'X-authenticated-user-token': accessToken,
                        'X-Channel-id': x_channel_id
                    }
                    responseImportSoluTemplateApi = requests.get(url=urlImportSoluTemplate,
                                                                 headers=headerImportSoluTemplateApi)
                    if responseImportSoluTemplateApi.status_code == 200:
                        responseImportSoluTemplateApi = responseImportSoluTemplateApi.json()
                        solutionIdSuc = responseImportSoluTemplateApi["result"]["solutionId"]
                        childsolutionid = solutionIdSuc
                        
                        urlSurveyProgramMapping = internal_kong_ip_survey + importsurveysolutiontoprogramurl + str(solutionIdSuc) + "?programId=" + programExternalId.lstrip().rstrip()
                        headeSurveyProgramMappingApi = {
                            'Authorization': authorization,
                            'X-authenticated-user-token': accessToken,
                            'X-Channel-id':x_channel_id
                        }
                        responseSurveyProgramMappingApi = requests.get(url=urlSurveyProgramMapping,headers=headeSurveyProgramMappingApi)
                        if responseSurveyProgramMappingApi.status_code == 200:
                            # print('Program Mapping Success')
                            surveyLink = None
                            solutionIdSuc = None
                            surveyExternalIdSuc = None
                            surveyLink = responseImportSoluTemplateApi["result"]["link"]
                            solutionIdSuc = responseImportSoluTemplateApi["result"]["solutionId"]
                            solutionExtIdSuc = responseImportSoluTemplateApi["result"]["solutionExternalId"]
        
                            return str(solutionIdSuc)
                        else:
                            print('Program Mapping Failed')
                    else:
                        print('Creating Child API Failed')
                else:
                    print('QuestionUploadApi Failed')

    def getQuestionUploadData(parentFolder):
        with open(parentFolder + '/questionUpload/uploadInternalIdsSheet.csv', 'r') as file:
            reader = csv.reader(file)
            data = list(reader)
        return data

    def preparesolutionUploadSheet(mainFilePath,parentFolder,solutionId):
        successSheetName = mainFilePath + "/SolutionFiles/" + solutionNameForSuccess + ".xlsx"
        # Load workbook
        shutil.copy(parentFolder + "user_input_file.xlsx", successSheetName)
        # Load workbook
        wb = load_workbook(parentFolder + "user_input_file.xlsx")

        # Read CSV data
        csv_data = Helpers.getQuestionUploadData(parentFolder)
        sheet_name = "uploadedQuestionIDs"
        # Check if the sheet already exists
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            # Find the next available row
            next_row = ws.max_row + 1
        else:
            # Create a new worksheet
            ws = wb.create_sheet(title=sheet_name)
            next_row = 1

        # Add CSV data to the worksheet
        for row in csv_data:
            for col_idx, cell_value in enumerate(row, start=1):
                ws.cell(row=next_row, column=col_idx, value=cell_value)
            next_row += 1

        # Save the workbook
        wb.save(successSheetName)
        wb = openpyxl.load_workbook(successSheetName)

        # Get the specified sheet or create it if it doesn't exist
        sheet_name = "details"
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.create_sheet(title=sheet_name)

        bg_color='00FF00'
        cell = ws.cell(row=2, column=6, value="solutionId")
        cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
        cell = ws.cell(row=3, column=6, value=solutionId)
        cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
        cell = ws.cell(row=2, column=7, value="programId")
        cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
        cell = ws.cell(row=3, column=7, value=programIdForSuccess)
        cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")

        # Save the workbook
        wb.save(successSheetName)
        return successSheetName

    def uploadSuccessSheetToBucket(solutionId,successSheetName,accessToken):
        persignedUrl = public_url_for_core_service + getpresignedurl
        successSheetExcel = successSheetName.split('/')[-1]
        presignedUrlBody = {
            "request": {
                solutionId: {
                    "files": [
                        successSheetExcel
                    ]
                }
            },
            "ref": "solution"
        }
        headerPreSignedUrl = {'Authorization': authorization,
                                   'X-authenticated-user-token': accessToken,
                                   'Content-Type': content_type}
        responseForPresignedUrl = requests.request("POST", persignedUrl, headers=headerPreSignedUrl,
                                                    data=json.dumps(presignedUrlBody))
        if responseForPresignedUrl.status_code == 200:
            presignedResponse = responseForPresignedUrl.json()
            programupdateData = presignedResponse['result']
            fileUploadUrl = presignedResponse['result'][solutionId]['files'][0]['url']
            if '?file=' in fileUploadUrl:
                downloadedurl = fileUploadUrl.split('?file=')[1]
            else:
                downloadedurl = None
            headers = {
                'Authorization': authorization,
                'X-authenticated-user-token': accessToken,
            }
            files={
                'file': open(successSheetName, 'rb')
            }
            response = requests.post(url=fileUploadUrl, headers=headers, files=files)
            if response.status_code == 200:
                # print("File Uploaded successfully")
                solutionFileData = programupdateData[solutionId]
                programUpdateDetails = {
                    "solutionId" : solutionId,
                    **solutionFileData
                }
                Helpers.getProgramDetailsMetaAndUpdate(programUpdateDetails,accessToken)
        return downloadSuccessSheet+downloadedurl


    def getProgramDetailsMetaAndUpdate(programMetaInfo,accessToken) :
        programMetaInfo['fileUploadedAt'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        programMetaInfo['fileUploadedBy'] = creatorId 
        programDetailsurl = internal_kong_ip_core + programdetailsapi + programIdForSuccess
        headerPreSignedUrl = {'Authorization': authorization,
                                'X-authenticated-user-token': accessToken,
                                'internal-access-token': internal_access_token,
                                'Content-Type': content_type}
        responseForProgramDetails = requests.request("GET", programDetailsurl, headers=headerPreSignedUrl)
        if responseForProgramDetails.status_code == 200:
            programdetailsres = responseForProgramDetails.json()
            if 'metaInformation' in programdetailsres['result']:
                metaInformation = programdetailsres['result']['metaInformation']

                # Check if 'successSheet' key is present in 'metaInformation'
                if 'uploadHistory' in metaInformation:
                    metaInformation['uploadHistory'].append(programMetaInfo)
                    Helpers.programUpdateMeta({"metaInformation":metaInformation},accessToken)
                else:
                    metaInformation['uploadHistory'] = [programMetaInfo]
                    Helpers.programUpdateMeta({"metaInformation":metaInformation},accessToken)
            else:
                Helpers.programUpdateMeta({"metaInformation":{"uploadHistory":[programMetaInfo]}},accessToken)

    def programUpdateMeta(body,accessToken):
        programUpdateUrl = internal_kong_ip_core + programupdateapi + programIdForSuccess
        headerPreSignedUrl = {'Authorization': authorization,
                                'X-authenticated-user-token': accessToken,
                                'internal-access-token': internal_access_token,
                                'Content-Type': content_type}
        responseForProgramUpdate = requests.request("POST", programUpdateUrl, headers=headerPreSignedUrl,
                                                    data=json.dumps(body))
        if responseForProgramUpdate.status_code == 200:
            return("program Updated successfully")

    def schedule_deletion(returnPathStr):
        def delete_file():
            try:
                time.sleep(15)
                if os.path.exists(returnPathStr):
                    if os.path.isfile(returnPathStr):
                        os.remove(returnPathStr)
                        print(f"File {returnPathStr} deleted successfully.")

                    elif os.path.isdir(returnPathStr):
                        shutil.rmtree(returnPathStr)
                else:
                    print(f"File {returnPathStr} not found.")
            except Exception as e:
                print(f"Error deleting file: {e}")

        threading.Thread(target=delete_file, daemon=True).start()


        
    def mainFunc(MainFilePath, addSolutionFile):
        global isProgramnamePresent

        parentFolder = Helpers.createFileStruct(MainFilePath, addSolutionFile)
        

        accessToken = Helpers.generateAccessToken(parentFolder)
        wbObservation = xlrd.open_workbook(addSolutionFile, on_demand=True)
        Helpers.SolutionFileCheck(addSolutionFile, accessToken, parentFolder, MainFilePath)
        surveyResp = Helpers.createSurveySolution(parentFolder, wbObservation, accessToken)
        surTempExtID = surveyResp[1]
        # surveyChildId = surveyResp[0]
        bodySolutionUpdate = {"status": "active", "isDeleted": False}
        Helpers.solutionUpdate(parentFolder, accessToken, surveyResp[0], bodySolutionUpdate)
        surveyChildId = Helpers.uploadSurveyQuestions(parentFolder, wbObservation, addSolutionFile, accessToken, surTempExtID,
                                surveyResp[0], millisecond)
        
        local = os.getcwd()
        sucessSheetName = Helpers.preparesolutionUploadSheet(MainFilePath,parentFolder,surveyChildId)
        clickheretodownload = Helpers.uploadSuccessSheetToBucket(surveyChildId,sucessSheetName,accessToken)
        Helpers.schedule_deletion(MainFilePath)
        return [surveyChildId,local+'/'+sucessSheetName,clickheretodownload]
    
    
    def loadSurveyFile(resourceFile):
        
        MainFilePath = Helpers.createFileStructForProgram(resourceFile)
        # print(MainFilePath,"return surveyResp[0]return surveyResp[0]")
        wbPgm = xlrd.open_workbook(resourceFile, on_demand=True)
        sheetNames = wbPgm.sheet_names()
        resourceLinkOrExtPGMcopy = '/'+str(resourceFile)
        if not os.path.isdir('InputFiles'):
            os.mkdir('InputFiles')
        shutil.copy(resourceLinkOrExtPGMcopy,'InputFiles' )

        solutionSL = Helpers.mainFunc(MainFilePath, os.path.join('InputFiles',resourceFile))
        return solutionSL
