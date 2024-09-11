#write your code
from flask import Flask, request , send_from_directory,jsonify
import os,time,sys
from dotenv import load_dotenv
import json 
import hashlib 
import jwt
from flask_cors import CORS
import numpy as np
import pymongo
import pandas as pd
import shutil
import openpyxl
from openpyxl import load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import PatternFill
from bson import json_util
# importing ObjectId from bson library
from bson.objectid import ObjectId
from datetime import datetime
import subprocess
# from backend.src.main.modules.helper import *




sys.path.append('../../..')
sys.path.append('../../../backend/src/main/modules/')
from backend.src.main.modules.xlsxObject import xlsxObject
from backend.src.main.modules.survey import SurveyCreate
from backend.src.main.modules.helper import Helpers
# from backend.src.main.modules.commom_config import config.ini
# from backend.src.main.modules import main


def myconverter(obj):
        if isinstance(obj, np.integer):
            return int(obj)
        elif isinstance(obj, np.floating):
            return float(obj)
        elif isinstance(obj, np.ndarray):
            return obj.tolist()
        elif isinstance(obj, datetime):
            return obj.__str__()


STATIC_PATH = os.path.join(os.getcwd(),"tmp")


app = Flask(__name__,static_url_path="/tmp/")

# enable CORS for the app 
CORS(app)
# get the base directory 
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
# load the .env file 
dotenv_path = os.path.join(BASE_DIR, '.env')  
# check env file  is accessible or not
if os.path.exists(dotenv_path):
    load_dotenv(dotenv_path)
else:
    print('".env" is missing.')
    sys.exit(1)
# connect to mongo db and collection instance function 
def connectDb(url,db,collection):
    client = pymongo.MongoClient(url)
    db = client[db]
    collectionData = db[collection]
    # print("collectionData",collectionData)
    return collectionData

def addComments(templatePath, errResponse):
    xlsxData = pd.read_excel(templatePath, sheet_name=None)
    errPath = templatePath.split(".")[0]+"_errFile"+".xlsx"
    shutil.copyfile(templatePath, errPath)

    workBook = load_workbook(errPath, data_only = True)
    try:
        for key in xlsxData.keys():
            newHeader = xlsxData[key].iloc[0]
            xlsxData[key] = xlsxData[key][1:]
            xlsxData[key].columns = newHeader
    except Exception as e:
        workBook.save(errPath)
        errResponse["result"]["errFileLink"] = os.environ.get("HOSTIP")+"/template/api/v1/errDownload?templatePath="+errPath
        return errResponse
    for result in errResponse["result"]:
        for errData in errResponse["result"][result]["data"]:
            if errData["columnName"] != "":
                try:
                    spreadSheet = workBook[errData["sheetName"]]
                except Exception as e:
                    print(e, errData["sheetName"]," sheet is missing")
                    continue
                try:
                    columnNumber = xlsxData[errData["sheetName"]].columns.get_loc(errData["columnName"])
                except Exception as e:
                    if spreadSheet.cell(2,1).comment is None:
                        spreadSheet.cell(2,1).comment=Comment("Error - "+errData["errMessage"]+"\n Suggestion -"+errData["suggestion"]+"\n" ,"admin")
                        spreadSheet.cell(2,1).fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type = "solid")
                    else:
                        spreadSheet.cell(2,1).comment=Comment(spreadSheet.cell(row=2, column=1).comment.text+"Error - "+errData["errMessage"]+"\n Suggestion -"+errData["suggestion"]+"\n" ,"admin")
                        spreadSheet.cell(2,1).fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type = "solid")
                    continue
                if type(errData["rowNumber"]) is list:
                    for rowIndex in errData["rowNumber"]:
                        if spreadSheet.cell(row=rowIndex+2, column=columnNumber+1).comment is None:
                            spreadSheet.cell(row=rowIndex+2, column=columnNumber+1).comment=Comment("Error - "+errData["errMessage"]+"\n Suggestion -"+errData["suggestion"]+"\n","admin")
                            spreadSheet.cell(row=rowIndex+2, column=columnNumber+1).fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type = "solid")
                        else:
                            spreadSheet.cell(row=rowIndex+2, column=columnNumber+1).comment=Comment(spreadSheet.cell(row=rowIndex+2, column=columnNumber+1).comment.text+"Error - "+errData["errMessage"]+"\n Suggestion -"+errData["suggestion"]+"\n","admin")
                            spreadSheet.cell(row=rowIndex+2, column=columnNumber+1).fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type = "solid")
                elif type(errData["rowNumber"]) is int:
                    if spreadSheet.cell(row=errData["rowNumber"]+2, column=columnNumber+1).comment is None:
                        spreadSheet.cell(row=errData["rowNumber"]+2, column=columnNumber+1).comment=Comment("Error - "+errData["errMessage"]+"\n Suggestion -"+errData["suggestion"]+"\n","admin")
                        spreadSheet.cell(row=errData["rowNumber"]+2, column=columnNumber+1).fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type = "solid")
                    else:
                        spreadSheet.cell(row=errData["rowNumber"]+2, column=columnNumber+1).comment=Comment(spreadSheet.cell(row=errData["rowNumber"]+2, column=columnNumber+1).comment.text+"Error - "+errData["errMessage"]+"\n Suggestion -"+errData["suggestion"]+"\n","admin")
                        spreadSheet.cell(row=errData["rowNumber"]+2, column=columnNumber+1).fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type = "solid") 
            else:
                if errData["errCode"] == 300:
                    workBook.create_sheet(errData["sheetName"])
                    try:
                        spreadSheet = workBook[errData["sheetName"]]
                    except Exception as e:
                        print(e, errData["sheetName"]," sheet is missing")
                        continue
                
                    spreadSheet.cell(2,1).comment=Comment("Error - "+errData["errMessage"]+"\n Suggestion -"+errData["suggestion"]+"\n" ,"admin")
                    spreadSheet.cell(2,1).fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type = "solid")
                    continue
                else:
                    try:
                        spreadSheet = workBook[errData["sheetName"]]
                    except Exception as e:
                        print(e, errData["sheetName"]," sheet is missing")
                        continue
                
                    for rowIndex in errData["rowNumber"]:
                        if spreadSheet.cell(rowIndex+2,1).comment is None: 
                            spreadSheet.cell(rowIndex+2,1).comment=Comment("Error - "+errData["errMessage"]+"\n Suggestion -"+errData["suggestion"]+"\n" ,"admin")
                            spreadSheet.cell(rowIndex+2,1).fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type = "solid")
                        else:
                            spreadSheet.cell(rowIndex+2,1).comment=Comment(spreadSheet.cell(rowIndex+2,1).comment.text+"Error - "+errData["errMessage"]+"\n Suggestion -"+errData["suggestion"]+"\n" ,"admin")
                            spreadSheet.cell(rowIndex+2,1).fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type = "solid")
                    continue

                    
    workBook.save(errPath)
    errResponse["result"]["errFileLink"] = os.environ.get("HOSTIP")+"/template/api/v1/errDownload?templatePath="+errPath
    return errResponse

# Login user API 
@app.route("/template/api/v1/authenticate", methods = ['POST'])
def login():
    req_body = request.get_json()
    try:
        # get the user name from request 
        userName = req_body['request']['email']
        # get the password from request and hash it in md5 
        password = hashlib.md5(req_body['request']['password'].encode('utf-8'))

        # connect to user collection 
        usersCollection = connectDb(os.environ.get('mongoURL'),os.environ.get('db'),'userCollection')
        
        # query the username and hashed password pair is present in DB
        users = usersCollection.count_documents({'userName' : userName , "password" : str(password.hexdigest())})

        # check the user result 
        if(users):
            # Exipry and other details can be added here
            message = {
                'iss': '',
                'email': userName
                }
            
            # secret key from the env file 
            signing_key = os.environ.get("SECRET_KEY")
            # encode the user name and expiry to create a token 
            try:
                encoded_jwt = jwt.encode({'message': message}, signing_key, algorithm='HS256')
            except Exception as e:
                encoded_jwt = ""
                print(e)

            # return the token after successful authentication 
            return {"status" : 200,"code" : "Authenticated","errorFlag" : False,"error" : [],"response" : {
                "accessToken" : encoded_jwt
            }}
        else:
            # return authentication failed error 
            return {"status" : 404,"code" : "Error","errorFlag" : True,"error" : ["Username / Password Doesn't Match"],"response" : {
                "accessToken" : "" }}
    except Exception as e:
        # throw the error 
        return {"status" : 500,"code" : str(e) ,"errorFlag" : True,"error" : ["Error in reaching server"],"response" : {
                "accessToken" : "" }}

# sign up API
@app.route("/template/api/v1/signup", methods = ['POST'])
def signup():
    req_body = request.get_json()
    # get the 'admin-token' from the request header 
    auth = request.headers.get('admin-token')
    # check for the auth token 
    if(not auth):
        # if the auth token is missing return authorization failed 
        return {"status" : 500,"code" : "Authorization Failed" , "result" : {"templateLinks" : ""}}
    else:
        # the auth token is present in the header and check the token present in the env file 
        if not auth == os.environ.get('admin-token'):
            return {"status" : 500,"code" : "Not Authorized" , "result" : {"templateLinks" : ""}}

    # if auth is checked 
    try:
        # get the username from request body 
        userName = req_body['request']['email']
        # get the password from request body and hash it
        password = hashlib.md5(req_body['request']['password'].encode('utf-8'))
        # connect to users collection in mongo DB
        usersCollection = connectDb(os.environ.get('mongoURL'),os.environ.get('db'),'userCollection')

        # get the current time 
        now = datetime.now()
        # query the given username
        users = usersCollection.count_documents({'userName' : userName})
        # check if the username is already present or not 
        if(users <= 0):
            # not present create the user in DB 
            users = usersCollection.insert_one({'userName' : userName , "password" : str(password.hexdigest()),"status" : "active","role" : "admin","createdAt" : str(now),"updatedAt" : str(now),"createdBy" : "admin"})
            # return success message 
            return {"status" : 200,"code" : "Authenticated","errorFlag" : False,"error" : [],"response" : "User created Successfully."}
        else:
            # return user already exists 
            return {"status" : 404,"code" : "Error","errorFlag" : True,"error" : ["UserName already exisiting."],"response" : {"accessToken" : "" }}
    except Exception as e:
        # return error 
        return {"status" : 500,"code" : str(e) ,"errorFlag" : True,"error" : ["Error in reaching server"],"response" : {"accessToken" : "" }}

# sample template downloader api
@app.route("/template/api/v1/download/sampleTemplate", methods = ['GET'])
def sample():
    errors = []

    # connect to db 
    try:       
        sampleTemplate = connectDb(os.environ.get("mongoURL"),os.environ.get("db"),os.environ.get("sampleTemplatesCollection"))
    except Exception as e:
        errors.append(e)
    templateListResp = []
    # find the list of entries from sampleTemplatesCollection
    sampleTemplateResponse = sampleTemplate.find({})
    for index in sampleTemplateResponse:
        templateListResp.append({"templateName" : index['templateName'], "templateLink" : index['templateLink'] , "templateCode" : index['templateCode']})

    return {"status" : 200,"code" : "OK" , "result" : {"templateLinks" : templateListResp}}


@app.route("/template/api/v1/add/sampleTemplate", methods = ['POST'])
def sampleAdd():
    errors = []
    # get body from the request 
    req_body = request.get_json()
    # fetch admin token from headers 
    admin_token = request.headers.get('admin-token')
    # validation of admin token 
    if(not admin_token):
        return {"status" : 500,"code" : "Admin Authorization key missing" , "result" : []}
    
    if not admin_token == os.environ.get("admin-token"):
        return {"status" : 500,"code" : "Admin Authorization Failed" , "result" : []}
    
    # connect with sampleTemplate collection 
    try:       
        sampleTemplate = connectDb(os.environ.get("mongoURL"),os.environ.get("db"),os.environ.get("sampleTemplatesCollection"))
    except Exception as e:
        errors.append(e)

    templateListResp = []
    # get the list of sample templates 
    sampleTemplateResponse = sampleTemplate.find({})
    maxTemplateCode = 0
    # get the max id value 
    for index in sampleTemplateResponse:
        if index['templateCode'] > maxTemplateCode :
            maxTemplateCode = index['templateCode']
    # increment one to max id value present in the DB 
    maxTemplateCode = maxTemplateCode +1
    # get current time 
    now = datetime.now()
 
    # Creating Dictionary of records to be inserted
    record = { "templateCode": maxTemplateCode,
              "templateName": req_body['request']['templateName'],
              "templateLink": req_body['request']['templateLink'],
              "createdBy": "admin",
              "updatedBy" : "admin",
              "createdAt" : now,
              "updatedAt" : now
              }
    
    db_result = sampleTemplate.insert_one(record)

    if not db_result.inserted_id : 
        errors.append("Insertion failed.")
        errors.append(db_result.inserted_id)
    

    templateListResp = []
    sampleTemplateResponse = sampleTemplate.find({})
    for index in sampleTemplateResponse:
        templateListResp.append({"templateName" : index['templateName'], "templateLink" : index['templateLink'] , "templateCode" : index['templateCode']})
    if len(errors) <= 0:
        return {"status" : 200,"code" : "OK" ,"message" : "Template added successfully", "result" : {"templateLinks" : templateListResp}}
    else:
        return {"status" : 200,"code" : "NOTOK" ,"message" : "Template adding failed", "result" : {}}

@app.route("/template/api/v1/update/sampleTemplate/<code>", methods = ['POST'])
def sampleUpdate(code):
    errors = []
    # get body from the request 
    req_body = request.get_json()
    # fetch admin token from headers 
    admin_token = request.headers.get('admin-token')
    # validation of admin token 
    if(not admin_token):
        return {"status" : 500,"code" : "Admin Authorization key missing" , "result" : []}
    
    if not admin_token == os.environ.get("admin-token"):
        return {"status" : 500,"code" : "Admin Authorization Failed" , "result" : []}
    
    # connect with sampleTemplate collection 
    try:       
        sampleTemplate = connectDb(os.environ.get("mongoURL"),os.environ.get("db"),os.environ.get("sampleTemplatesCollection"))
    except Exception as e:
        errors.append(e)

    # get current time 
    now = datetime.now()

    argKeys = [key for key in req_body['request'].keys()]

    if len(argKeys) <= 0:
        return {"status" : 200,"code" : "NOTOK" ,"message" : "Body cannot be empty", "result" : {}}
    # Creating Dictionary of records to be updated
    record = { 
              "updatedAt" : now
              }
    # identify the key updating in the body 
    if 'templateName' in argKeys :
        record["templateName"] = req_body['request']['templateName']
    if  'templateLink' in argKeys :
        record["templateLink"] = req_body['request']['templateLink']

    
    # create the update value 
    newvalues = { "$set":  record  }
    # filter by template code passed in the api 
    filter = { 'templateCode': int(code) }
 
    updateResult = sampleTemplate.update_one(filter, newvalues)

    templateListResp = []
    sampleTemplateResponse = sampleTemplate.find({})
    for index in sampleTemplateResponse:
        templateListResp.append({"templateName" : index['templateName'], "templateLink" : index['templateLink'] , "templateCode" : index['templateCode']})
    
    if updateResult.matched_count > 0 :
        return {
            "code": "OK",
            "count": updateResult.matched_count,
            "error": errors,
            "message" : "Sample tempalate Updated Successfully",
            "result": { "templateLinks" : templateListResp}
            }
    else:
        return {
            "code": "OK",
            "count": updateResult.matched_count,
            "message" : "Sample tempalate Update failed",
            "error": errors,
            "result": {}
                
            }


# API to upload excel file to server 
@app.route("/template/api/v1/upload", methods = ['POST'])
def upload():

    # get auth Token for validation
    # auth = request.headers.get('Authorization')
    # # get SECRET_KEY for validation
    # signing_key = os.environ.get("SECRET_KEY")

    # payload = False
    # # check if auth token is present in the header 
    # if(not auth):
    #     return {"status" : 500,"code" : "Authorization Failed" , "result" : {"templateLinks" : ""}}
    # else:

    #     # decode the payload with signing_key to check if the user is authentic 
    #     # print("=-=-=-==-=-> ",auth)
    #     payload = jwt.decode(auth, signing_key, algorithms=['HS256'])

    # if(not payload):
    #     return {"status" : 500,"code" : "Authorization Failed" , "result" : {"templateLinks" : "True"}}
    
    # set the allowed extensions to upload 
    ALLOWED_EXTENSIONS = set(['xlsx'])

    # set static path to upload to the file  create the folder in server if not created.
    if not os.path.exists(STATIC_PATH):
        os.makedirs(STATIC_PATH)
    # check the request method 
    if request.method == 'POST':
        # check if the post request has the file part
        if 'file' not in request.files:
            return {"status" : 500,"code" : "Required key missing!" , "result" : {"templateLinks" : ""}}
        # get the file from request 
        file = request.files['file']
        # extract extension from file     
        ext = file.filename.split('.')
        if file and ext[1] in ALLOWED_EXTENSIONS:
            filename = file.filename
            #fileName clearing.
            filename = filename.replace(" ","_")
            filenameArr = filename.split(".")

            # ts stores the time in seconds
            ts = str(time.time()).replace(".","-")
            # checnge the file name 
            finalFileName = str(filenameArr[0])+str(ts)+"."+str(filenameArr[1])
            try:
                # save the file in the server 
                file.save(os.path.join(STATIC_PATH, finalFileName))
            except Exception as e:
                # print any error in saving file 
                print(e)
                # return error status 
                return {"status" : 500,"code" : "Server Error" , "result" : {"templatePath" : ""}}
            # return path of the file saved in the local server 
            return {"status" : 200,"code" : "OK" , "result" : {"templatePath" : os.path.join(STATIC_PATH, finalFileName),"templateName" : finalFileName}}

        
        return {"status" : 404,"code" : "File Error." , "result" : {"templateLinks" : ""}}
        
@app.route("/template/api/v1/validate", methods = ['POST'])
def validate():
    req_body = request.get_json()
    templateFolderPath = req_body["request"]["templatePath"]
    templateCode = req_body["request"]["templateCode"]

    # Token validation
    # auth = request.headers.get("Authorization")
    # signing_key = os.environ.get("SECRET_KEY")
    # payload = False
    # if(not auth):
    #     return {"status" : 500,"code" : "Authorization Failed" , "result" : {"templateLinks" : ""}}
    # else:
    #     try:
    #         payload = jwt.decode(auth, signing_key, algorithms=['HS256'])
    #     except Exception as e:
    #         print(e)

    # if(not payload):
    #     return {"status" : 500,"code" : "Authorization Failed" , "result" : {"templateLinks" : "True"}}
    

    basicErrors = xlsxObject(templateCode, templateFolderPath)
    print
    # main

    if basicErrors.success:
        valErr = basicErrors.basicCondition()
        advValErr = basicErrors.customCondition()
        return addComments(templateFolderPath,{"status" : 200,"code" : "OK" , "result" : {"basicErrors" : valErr,"advancedErrors" : advValErr}})
    else:
        return {"status" : 404,"code" : "ERROR" , "result" :{},"message":"Please check template id"}


@app.route("/template/api/v1/errDownload", methods = ['GET'])
def errDownload():
    templateFolderPath = request.args.get("templatePath")
    return send_from_directory(os.path.dirname(templateFolderPath), os.path.basename(templateFolderPath), as_attachment=True)

# show the user roles list 
@app.route("/template/api/v1/userRoles/list", methods = ['GET'])
def userRoles():
    returnResponse = {}
    # connect to conditions Collection
    subRoles = connectDb(os.environ.get('mongoURL'),os.environ.get('db'),os.environ.get('conditionsCollection'))
    # query recommendedForCheck condition 
    returnResponseTmp = subRoles.find({"name" : "recommendedForCheck"})
    
    if returnResponseTmp:
        returnResponse["status"] = 200
        returnResponse["code"] = "OK"
        # return roles details 
        returnResponse["result"] = returnResponseTmp[0]['recommendedForCheck']['roles']
    return returnResponse


# Update and add new subroles using this API
@app.route("/template/api/v1/userRoles/update", methods = ['POST'])
def update():

    error = []
    result = {}

    req_body = request.get_json()
    auth = request.headers.get('admin-token')
    request["auth"] = auth


    # Auth code check
    if(not auth):
        return {"status" : 500,"code" : "Authorization Failed" , "result" : []}
    else:
        if not auth == os.environ.get("admin-token"):
            return {"status" : 500,"code" : "Not Authorized" , "result" : []}
    try:
        mydict = {}

        if req_body["request"]["code"] == None or req_body["request"]["title"] == None or req_body["request"]["code"] == "" or req_body["request"]["title"] == "" or req_body["request"]["_id"] == None or req_body["request"]["_id"] == "" :
            error.append("Required value missing")
        else:
            subRoles = connectDb(os.environ.get("mongoURL"),os.environ.get("db"),os.environ.get("conditionsCollection"))
            returnResponseTmp = subRoles.find({"name" : "recommendedForCheck"})
    
            mydict = {"code" : req_body["request"]["code"] , "title" : req_body["request"]["title"], "_id" : req_body["request"]["_id"]}
            chechSubRole = 0
            codeFlag = False
            idFlag = False

            currentSubRoles = returnResponseTmp[0]["recommendedForCheck"]["roles"]
            for index in currentSubRoles:
                if index["code"] == req_body["request"]["code"]:
                    codeFlag = True
                if index["_id"] == req_body["request"]["_id"]:
                    idFlag = True

            if codeFlag:
                error.append("Duplicate Code error")

            # update Subrole 
            if idFlag :
                indexCount = 0
                for index in currentSubRoles:
                    if index["_id"] == req_body["request"]["_id"]:
                        currentSubRoles[indexCount]["code"] = req_body["request"]["code"]
                        currentSubRoles[indexCount]["title"] = req_body["request"]["title"]
                    indexCount += 1
                findQuery = { "name" : "recommendedForCheck" }
                newvalues = { "$set": { "recommendedForCheck.roles": currentSubRoles } }

                subRoles.update_one(findQuery, newvalues)
                result = {
                    "message" : "subRoles updated successfully.",
                    "_id" : req_body["request"]["_id"],
                    "title" : req_body["request"]["title"],
                    "code" : req_body["request"]["code"]
                }
                # Add new subroles   
            elif len(error) <= 0:

                currentSubRoles.append(mydict)
                findQuery = { "name" : "recommendedForCheck" }
                newvalues = { "$set": { "recommendedForCheck.roles": currentSubRoles } }

                subRoles.update_one(findQuery, newvalues)
                result = {
                    "message" : "subRoles added successfully."
                }
    except Exception as e:
        print(e)
        error = "Key missing."

    
    return {"status" : 200,"code" : "OK", "result" : result,"error" : error}

# list the validation rules 
@app.route("/template/api/v1/validations/list", methods = ['GET'])
def listValidations():
    client = pymongo.MongoClient(os.environ.get('mongoURL'))
    args = request.args
    validationsRes = None
    # Token validation
    admin_token = request.headers.get("admin-token")

    if(not admin_token):
        return {"status" : 500,"code" : "Admin Authorization key missing" , "result" : []}
    
    if not admin_token == os.environ.get("admin-token"):
        return {"status" : 500,"code" : "Admin Authorization Failed" , "result" : []}
    
    # fetching the keys from arguments 
    argKeys = [key for key in args.keys()]
    query = {}
    errors = []
    # preparing the query to find() from the keys passed in arguments 
    if "id" in argKeys:
        query.update({"id": args["id"]})
    if "resourceType" in argKeys:
        query.update({"resourceType": args["resourceType"]})
    
    # connecting with DB and validations collection 
    try:       
        validationsCollection = connectDb(os.environ.get("mongoURL"),os.environ.get("db"),os.environ.get("validationsCollection"))
    except Exception as e:
        errors.append(e)

    # running the find query 
    try:  
        validationsRes = validationsCollection.find(query)
    except Exception as e:
        print(e)
    # get count of result 
    try:
        validationsCount = validationsCollection.count_documents(query)
    except Exception as e:
        print(e)
    result = []

    # preare result to display 
    for index in validationsRes:
        index["_id"] = str(index["_id"])
        result.append(index)

    return {"status" : 200,"code" : "OK","count" : validationsCount, "result" : json.loads(json_util.dumps(result)),"error" : errors}

# update validation using id 
@app.route("/template/api/v1/validations/update/<_id>", methods = ['POST'])
def updateValidations(_id):
    errors = []
    req_body = request.get_json()

    # Token validation
    admin_token = request.headers.get("admin-token")

    if(not admin_token):
        return {"status" : 500,"code" : "Admin Authorization key missing" , "result" : []}

    if not admin_token == os.environ.get("admin-token"):
        return {"status" : 500,"code" : "Admin Authorization Failed" , "result" : []}
    # check if the body have validations key in it 
    try:
        validations = req_body['validations']
    except Exception as e:
        # throw error if the key is missing 
        errors.append(str(e)) 
        return {
            "status" : 200,
            "code": "NOTOK",
            "count": 0,
            "error": errors,
            "result": [
                {
                "message" : "Key missing in request body"
                }
                ]
            }


    try: 
        # selecting the validation based on Id
        filter = { '_id': ObjectId(_id) }

    except Exception as e:
        errors.append(str(e))
        print(errors)
        return {
            "code": "OK",
            "count": 0,
            "error": errors,
            "result": []
            }

    
    # Values to be updated.
    newvalues = { "$set": { 'validations': validations } }
    
    try:   
        validationsCollection = connectDb(os.environ.get("mongoURL"),os.environ.get("db"),os.environ.get("validationsCollection"))
    except Exception as e:
        errors.append(e)
        
    try:
        # Using update_one() method for single updation.
        updateResult = validationsCollection.update_one(filter, newvalues)
    except Exception as e:
        updateResult = {}
        updateResult['matched_count'] = 0
        errors.append(e)
    
    # return response 
    if updateResult.matched_count > 0 :
        return {
            "code": "OK",
            "count": updateResult.matched_count,
            "error": errors,
            "result": [
                {
                "message" : "Updated successfully"
                }
                ]
            }
    else:
        return {
            "code": "OK",
            "count": updateResult.matched_count,
            "error": errors,
            "result": [
                {
                "message" : "Updated failed"
                }
                ]
            }


# list of condition rules
@app.route("/template/api/v1/conditions/list", methods = ['GET'])
def listConditions():
    client = pymongo.MongoClient(os.environ.get('mongoURL'))
    args = request.args
    validationsRes = None
    # Token validation
    admin_token = request.headers.get("admin-token")

    if(not admin_token):
        return {"status" : 500,"code" : "Admin Authorization key missing" , "result" : []}
    
    if not admin_token == os.environ.get("admin-token"):
        return {"status" : 500,"code" : "Admin Authorization Failed" , "result" : []}
    
    # fetching the keys from arguments 
    argKeys = [key for key in args.keys()]
    query = {}
    errors = []
    # preparing the query to find() from the keys passed in arguments 
    if "id" in argKeys:
        query.update({"id": args["id"]})
    if "name" in argKeys:
        query.update({"name": args["name"]})
    
    
    # connecting with DB and conditions collection 
    try:       
        conditionsCollection = connectDb(os.environ.get("mongoURL"),os.environ.get("db"),os.environ.get("conditionsCollection"))
    except Exception as e:
        errors.append(e)

    # running the find query 
    try:  
        conditionsRes = conditionsCollection.find(query)
    except Exception as e:
        print(e)
    # get count of result 
    try:
        conditionsCount = conditionsCollection.count_documents(query)
    except Exception as e:
        print(e)
    result = []

    # preare result to display 
    for index in conditionsRes:
        index["_id"] = str(index["_id"])
        result.append(index)

    return {"status" : 200,"code" : "OK","count" : conditionsCount, "result" : json.loads(json_util.dumps(result)),"error" : errors}



@app.route("/template/api/v1/conditions/update/<_id>", methods=['POST'])
def update_conditions(_id):
    try:
        errors = []
        update_fields = request.get_json()

        admin_token = request.headers.get("admin-token")

        if not admin_token:
            return jsonify({"status": 500, "code": "Admin Authorization key missing", "result": []})

        if not admin_token == os.environ.get("admin-token"):
            return jsonify({"status": 500, "code": "Admin Authorization Failed", "result": []})

        try:
            filter = {'_id': ObjectId(_id)}
        except Exception as e:
            errors.append(str(e))
            return jsonify({"code": "OK", "count": 0, "error": errors, "result": []})

        if not update_fields:
            return jsonify({"code": "OK", "count": 0, "error": errors, "result": [{"message": "No fields to update"}]})

        # Defining the  fields that should not be updated
        restricted_fields = ["_id", "name"]

        # Check if any restricted field is being updated
        for restricted_field in restricted_fields:
            if restricted_field in update_fields:
                errors.append(f"Cannot update restricted field '{restricted_field}'")
                return jsonify({"code": "NOTOK", "count": 0, "error": errors, "result": [{"message": "You are not allowed to update this field "}]})

        update_query = {}
        for key, value in update_fields.items():
            update_query["$set"] = {key: value}

        try:
            conditionscollection = connectDb(os.environ.get("mongoURL"), os.environ.get("db"),
                                               os.environ.get("conditionsCollection"))
            
            
            update_result = None
            try:
                update_result = conditionscollection.update_one(filter, update_query)
            except Exception as e:
                errors.append(str(e))
            
            if update_result and update_result.modified_count > 0:
                return jsonify({"code": "OK", "count": update_result.modified_count, "error": errors, "result": [{"message": "Updated successfully"}]})
            else:
                return jsonify({"code": "NOTOK", "count": 0, "error": errors, "result": [{"message": "No documents were updated"}]})
        except Exception as e:
            errors.append(str(e))
            return jsonify({"code": "NOTOK", "count": 0, "error": errors, "result": [{"message": "Error during update"}]})

    except Exception as e:
        # Handle unexpected exceptions and return a generic error message
        return jsonify({"status": 500, "code": "Internal Server Error", "result": [{"message": "An error occurred"}]})
    

@app.route('/template/api/v1/survey/getSolutions', methods=['POST'])
def fetchSurveySolutions():
    resurceType = request.get_json()
    # Token validation
    # auth = request.headers.get("Authorization")
    # signing_key = os.environ.get("SECRET_KEY")
    # payload = False
    # if(not auth):
    #     return {"status" : 500,"code" : "Authorization Failed" , "result" : {"templateLinks" : ""}}
    # else:
    #     try:
    #         payload = jwt.decode(auth, signing_key, algorithms=['HS256'])
    #     except Exception as e:
    #         print(e)

    # if(not payload):
    #     return {"status" : 500,"code" : "Authorization Failed" , "result" : {"templateLinks" : "True"}}

    survey = SurveyCreate()
    access_token = survey.generate_access_token()
    fetchedSolutionList=survey.fetch_solution_id(access_token,resurceType['resourceType'])

    if fetchedSolutionList:
        return jsonify({"status": 200, "code": "Success","SolutionList":fetchedSolutionList})
    
    else:
        return jsonify({"status": 400, "code": "NOTOK","SolutionList":"Error in getting the list of solutions"})



@app.route('/template/api/v1/survey/downloadSolutions', methods=['POST'])
def fetchSurveySolutions_Csv():
    resurceType = request.get_json()

    # Token validation
    # auth = request.headers.get("Authorization")
    # signing_key = os.environ.get("SECRET_KEY")
    # payload = False
    # if(not auth):
    #     return {"status" : 500,"code" : "Authorization Failed" , "result" : {"templateLinks" : ""}}
    # else:
    #     try:
    #         payload = jwt.decode(auth, signing_key, algorithms=['HS256'])
    #     except Exception as e:
    #         print(e)

    # if(not payload):
    #     return {"status" : 500,"code" : "Authorization Failed" , "result" : {"templateLinks" : "True"}}

    survey = SurveyCreate()
    access_token = survey.generate_access_token()
    csvFilePath=survey.fetch_solution_id_csv(access_token,resurceType['resourceType'])

    if csvFilePath:
        return jsonify({"status": 200, "code": "Success","csvFilePath":csvFilePath})
    
    else:
        return jsonify({"status": 400, "code": "NOTOK","SolutionList":"Error in getting the list of solutions"})


@app.route('/template/api/v1/survey/create', methods=['POST'])
def create():
    req = request.get_json()
    helperInstance = Helpers
    resourceFile=helperInstance.loadSurveyFile(req['file'])
    # Token validation
    # auth = request.headers.get("Authorization")
    # signing_key = os.environ.get("SECRET_KEY")
    # payload = False
    # if(not auth):
    #     return {"status" : 500,"code" : "Authorization Failed" , "result" : {"templateLinks" : ""}}
    # else:
    #     try:
    #         payload = jwt.decode(auth, signing_key, algorithms=['HS256'])
    #     except Exception as e:
    #         print(e)

    # if(not payload):
    #     return {"status" : 500,"code" : "Authorization Failed" , "result" : {"templateLinks" : "True"}}

    if resourceFile:
        return jsonify({"status": 200, "code": "Success", "result": [{"solutionId":resourceFile[0],"successSheet":resourceFile[1],"downloadbleUrl":resourceFile[2]}]})
    else :
        return jsonify({"status": 500, "code": "NOTOK","massege":"Could not create survey solution"})
    
if (__name__ == '__main__'):
    app.run(host=os.environ.get("HOSTIP")  , port=os.environ.get("FLASK_RUN_PORT") , debug=True)
    